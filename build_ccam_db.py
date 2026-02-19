"""
T2A Assistant — CCAM Database Builder
Parses ATIH CCAM data files and creates a structured SQLite database
with full-text search capabilities.
"""

import sqlite3
import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(r"C:\Users\Sammy\Work\t2a-assistant\data\ccam")
DB_PATH = Path(r"C:\Users\Sammy\Work\t2a-assistant\data\ccam.db")

# Column indices in CCAM_Final_2026 sheet (0-based)
COL = {
    "subdivision": 0,       # Subdivision (titres) / Code à 7 caractères
    "extension_pmsi": 1,    # Extension PMSI
    "code_ext": 2,          # Code à 7 caractères (et extension PMSI)
    "code_all": 3,          # Subdivision / Code toutes lignes
    "code_sub": 4,          # Subdivision / Code à 7 caractères (et extension PMSI) / modificateurs
    "text": 5,              # Texte : titre-libellés-notes
    "has_info": 6,          # Informations complémentaires avis HAS
    "coding_instruction": 7,# Consigne de codage pour le PMSI
    "mod_type": 8,          # Type de Modification
    "version": 9,           # Version CCAM_AM/PMSI concernée
    "line_type": 10,        # Typo ligne libellé
    "activity": 11,         # Code Activité
    "phase": 12,            # Phase
    "rc": 13,               # RC
    "ap": 14,               # AP
    "etm": 15,              # ETM
    "rgt": 16,              # Rgt
    "classant": 17,         # Caractère classant
    "fg2024_p0": 18,        # Profil FG 2024 Phase 0
    "fg2024_p1": 19,        # Profil FG 2024 Phase 1
    "fg2024_p2": 20,        # Profil FG 2024 Phase 2
    "fg2024_p3": 21,        # Profil FG 2024 Phase 3
    "fg2023_p0": 22,        # Profil FG 2023 Phase 0
    "fg2023_p1": 23,        # Profil FG 2023 Phase 1
    "fg2023_p2": 24,        # Profil FG 2023 Phase 2
    "fg2023_p3": 25,        # Profil FG 2023 Phase 3
    "ffm_se_2020": 26,      # FFM ; SE1-SE4 ; 2020
    "ffm_se_2019": 27,      # FFM ; SE1-SE6 ; 2019
    "icr_public": 28,       # ICR "public" activité 1
    "icr_private": 29,      # ICR "privé" activité 1
    "icr_act4": 30,         # ICR activité 4
    "icr_anapath": 31,      # ICR anapath
    "icr_rea": 32,          # ICR réa
    "modifiers": 33,        # Modificateurs activités 1/2/3/4/5
    "gestures_text": 34,    # Gestes complémentaires texte
    "gestures_act123": 35,  # Gestes complémentaires activité 1/2/3 (base)
    "gestures_act4": 36,    # Gestes complémentaires activité 4 (base)
    "gestures_act5": 37,    # Gestes complémentaires activité 5 (base)
    "anesthesia": 38,       # Anesthésies complémentaires
    "construction_note": 39,# Note de construction
    "denomination": 40,     # Dénomination phase/activité/anesthésie/CEC
    "has_date": 41,         # Date avis ANAES/HAS
    "chap_num": 42,         # N° 1e subdivision : chapitre
    "chap_title": 43,       # Titre 1e subdivision : chapitre
    "subchap_num": 44,      # N° 2e subdivision : sous-chapitre
    "subchap_title": 45,    # Titre 2e subdivision : sous-chapitre
    "para_num": 46,         # N° 3e subdivision : paragraphe
    "para_title": 47,       # Titre 3e subdivision : paragraphe
    "subpara_num": 48,      # N° 4e subdivision : sous-paragraphe
    "subpara_title": 49,    # Titre 4e subdivision : sous-paragraphe
    "date_start": 50,       # Date de début de validité
    "date_end": 51,         # Date de fin de validité
}

# Regex for CCAM code: 4 letters + 3 digits (7 chars)
CCAM_CODE_RE = re.compile(r'^[A-Z]{4}\d{3}$')


def strip_accents(text):
    """Remove accents for search normalization."""
    if not text:
        return ""
    nfkd = unicodedata.normalize('NFKD', str(text))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_for_search(text):
    """Normalize text for FTS indexing."""
    if not text:
        return ""
    text = strip_accents(str(text).lower())
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def safe_str(val):
    """Convert value to string, handling None."""
    if val is None:
        return None
    return str(val).strip() if str(val).strip() else None


def safe_float(val):
    """Convert value to float, handling None and non-numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def create_database():
    """Create the SQLite database schema."""
    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()

    # Main CCAM codes table
    c.execute("""
        CREATE TABLE ccam_codes (
            code TEXT PRIMARY KEY,
            extension_pmsi TEXT,
            code_full TEXT,
            label TEXT NOT NULL,
            label_normalized TEXT,
            coding_instruction TEXT,
            activity TEXT,
            phase TEXT,
            classant TEXT,
            fg2024_p0 TEXT,
            fg2024_p1 TEXT,
            fg2024_p2 TEXT,
            fg2024_p3 TEXT,
            icr_public REAL,
            icr_private REAL,
            icr_act4 REAL,
            icr_anapath REAL,
            icr_rea REAL,
            modifiers TEXT,
            gestures_text TEXT,
            gestures_act123 TEXT,
            gestures_act4 TEXT,
            gestures_act5 TEXT,
            anesthesia TEXT,
            construction_note TEXT,
            chapter_num TEXT,
            chapter_title TEXT,
            subchapter_num TEXT,
            subchapter_title TEXT,
            paragraph_num TEXT,
            paragraph_title TEXT,
            subparagraph_num TEXT,
            subparagraph_title TEXT,
            date_start TEXT,
            date_end TEXT,
            has_info TEXT,
            has_date TEXT,
            ffm_se TEXT
        )
    """)

    # Chapter hierarchy
    c.execute("""
        CREATE TABLE chapters (
            num TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            level INTEGER NOT NULL,
            parent_num TEXT
        )
    """)

    # Complementary gestures (associations)
    c.execute("""
        CREATE TABLE associations (
            code TEXT NOT NULL,
            associated_code TEXT NOT NULL,
            association_type TEXT,
            activity TEXT,
            PRIMARY KEY (code, associated_code, activity)
        )
    """)

    # FTS5 virtual table for full-text search
    c.execute("""
        CREATE VIRTUAL TABLE ccam_fts USING fts5(
            code,
            label,
            label_normalized,
            chapter_title,
            subchapter_title,
            paragraph_title,
            coding_instruction,
            content='ccam_codes',
            content_rowid='rowid'
        )
    """)

    conn.commit()
    return conn


def parse_complementary_file(conn):
    """Parse the fichier complémentaire Excel and populate the database."""
    print("Loading fichier complémentaire...")
    wb = load_workbook(str(DATA_DIR / "fichier_complementaire_ccam_2025_v4.xlsx"), read_only=True)
    ws = wb["CCAM_Final_2026"]

    c = conn.cursor()
    codes_inserted = 0
    associations_inserted = 0
    chapters_seen = set()
    current_chapter = {}

    rows = list(ws.iter_rows(values_only=True))
    print(f"  Total rows: {len(rows)}")

    # Skip header row
    for i, row in enumerate(rows[1:], start=2):
        if len(row) < 52:
            continue

        line_type = safe_str(row[COL["line_type"]])
        subdivision = safe_str(row[COL["subdivision"]])
        code_sub = safe_str(row[COL["code_sub"]])
        text = safe_str(row[COL["text"]])

        # Track chapter hierarchy
        chap_num = safe_str(row[COL["chap_num"]])
        chap_title = safe_str(row[COL["chap_title"]])
        subchap_num = safe_str(row[COL["subchap_num"]])
        subchap_title = safe_str(row[COL["subchap_title"]])
        para_num = safe_str(row[COL["para_num"]])
        para_title = safe_str(row[COL["para_title"]])
        subpara_num = safe_str(row[COL["subpara_num"]])
        subpara_title = safe_str(row[COL["subpara_title"]])

        # Update current chapter context
        if chap_num:
            current_chapter["chap_num"] = chap_num
            current_chapter["chap_title"] = chap_title
        if subchap_num:
            current_chapter["subchap_num"] = subchap_num
            current_chapter["subchap_title"] = subchap_title
        if para_num:
            current_chapter["para_num"] = para_num
            current_chapter["para_title"] = para_title
        if subpara_num:
            current_chapter["subpara_num"] = subpara_num
            current_chapter["subpara_title"] = subpara_title

        # Insert chapters
        for level, (num_key, title_key) in enumerate([
            ("chap_num", "chap_title"),
            ("subchap_num", "subchap_title"),
            ("para_num", "para_title"),
            ("subpara_num", "subpara_title"),
        ]):
            n = safe_str(row[COL[num_key.replace("chap", "chap").replace("subchap", "subchap").replace("para", "para").replace("subpara", "subpara")]])
            # Use the COL mapping
            col_num_key = {"chap_num": "chap_num", "subchap_num": "subchap_num", "para_num": "para_num", "subpara_num": "subpara_num"}[num_key]
            col_title_key = {"chap_title": "chap_title", "subchap_title": "subchap_title", "para_title": "para_title", "subpara_title": "subpara_title"}[title_key]
            num_val = safe_str(row[COL[col_num_key]])
            title_val = safe_str(row[COL[col_title_key]])
            if num_val and title_val and num_val not in chapters_seen:
                chapters_seen.add(num_val)
                parent = None
                if level == 1:
                    parent = current_chapter.get("chap_num")
                elif level == 2:
                    parent = current_chapter.get("subchap_num")
                elif level == 3:
                    parent = current_chapter.get("para_num")
                try:
                    c.execute("INSERT OR IGNORE INTO chapters (num, title, level, parent_num) VALUES (?, ?, ?, ?)",
                              (num_val, title_val, level, parent))
                except:
                    pass

        # Only process actual CCAM code lines (7 chars, 4 letters + 3 digits)
        if not code_sub or not CCAM_CODE_RE.match(code_sub):
            continue

        if not text or line_type not in ('L', 'LV', None):
            # L = libellé (label line), LV = label line with validity
            # Some codes might have line_type as None
            if line_type in ('T', 'NT', 'M'):
                continue

        code = code_sub
        label = text
        if not label:
            continue

        extension = safe_str(row[COL["extension_pmsi"]])
        code_full = safe_str(row[COL["code_ext"]]) or code

        # Extract all fields
        try:
            c.execute("""
                INSERT OR REPLACE INTO ccam_codes VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
            """, (
                code,
                extension,
                code_full,
                label,
                normalize_for_search(label),
                safe_str(row[COL["coding_instruction"]]),
                safe_str(row[COL["activity"]]),
                safe_str(row[COL["phase"]]),
                safe_str(row[COL["classant"]]),
                safe_str(row[COL["fg2024_p0"]]),
                safe_str(row[COL["fg2024_p1"]]),
                safe_str(row[COL["fg2024_p2"]]),
                safe_str(row[COL["fg2024_p3"]]),
                safe_float(row[COL["icr_public"]]),
                safe_float(row[COL["icr_private"]]),
                safe_float(row[COL["icr_act4"]]),
                safe_float(row[COL["icr_anapath"]]),
                safe_float(row[COL["icr_rea"]]),
                safe_str(row[COL["modifiers"]]),
                safe_str(row[COL["gestures_text"]]),
                safe_str(row[COL["gestures_act123"]]),
                safe_str(row[COL["gestures_act4"]]),
                safe_str(row[COL["gestures_act5"]]),
                safe_str(row[COL["anesthesia"]]),
                safe_str(row[COL["construction_note"]]),
                current_chapter.get("chap_num"),
                current_chapter.get("chap_title"),
                current_chapter.get("subchap_num"),
                current_chapter.get("subchap_title"),
                current_chapter.get("para_num"),
                current_chapter.get("para_title"),
                current_chapter.get("subpara_num"),
                current_chapter.get("subpara_title"),
                safe_str(row[COL["date_start"]]),
                safe_str(row[COL["date_end"]]),
                safe_str(row[COL["has_info"]]),
                safe_str(row[COL["has_date"]]),
                safe_str(row[COL["ffm_se_2020"]]),
            ))
            codes_inserted += 1
        except Exception as e:
            if codes_inserted < 5:
                print(f"  [WARN] Row {i}: {e}")

        # Parse associations from gestures columns (text + structured)
        for gesture_col in ["gestures_text", "gestures_act123", "gestures_act4", "gestures_act5"]:
            gesture_val = safe_str(row[COL[gesture_col]])
            if gesture_val:
                # Extract CCAM codes from gesture text (4 letters + 3 digits pattern)
                associated_codes = re.findall(r'[A-Z]{4}\d{3}', gesture_val)
                activity = gesture_col.replace("gestures_", "")
                for assoc_code in associated_codes:
                    if assoc_code != code:
                        try:
                            c.execute("""
                                INSERT OR IGNORE INTO associations (code, associated_code, association_type, activity)
                                VALUES (?, ?, 'complementary_gesture', ?)
                            """, (code, assoc_code, activity))
                            associations_inserted += 1
                        except:
                            pass

        # Parse anesthesia associations
        anesthesia_val = safe_str(row[COL["anesthesia"]])
        if anesthesia_val:
            anesthesia_codes = re.findall(r'[A-Z]{4}\d{3}', anesthesia_val)
            for anest_code in anesthesia_codes:
                if anest_code != code:
                    try:
                        c.execute("""
                            INSERT OR IGNORE INTO associations (code, associated_code, association_type, activity)
                            VALUES (?, ?, 'complementary_anesthesia', 'anesthesia')
                        """, (code, anest_code))
                        associations_inserted += 1
                    except:
                        pass

    wb.close()
    conn.commit()
    print(f"  Codes inserted: {codes_inserted}")
    print(f"  Associations inserted: {associations_inserted}")
    print(f"  Chapters inserted: {len(chapters_seen)}")
    return codes_inserted


def build_fts_index(conn):
    """Populate the FTS5 full-text search index."""
    print("Building FTS index...")
    c = conn.cursor()
    c.execute("""
        INSERT INTO ccam_fts (rowid, code, label, label_normalized, chapter_title, subchapter_title, paragraph_title, coding_instruction)
        SELECT rowid, code, label, label_normalized, chapter_title, subchapter_title, paragraph_title, coding_instruction
        FROM ccam_codes
    """)
    conn.commit()
    print("  FTS index built.")


def create_indexes(conn):
    """Create additional indexes for fast lookups."""
    print("Creating indexes...")
    c = conn.cursor()
    c.execute("CREATE INDEX IF NOT EXISTS idx_codes_chapter ON ccam_codes(chapter_num)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_codes_subchapter ON ccam_codes(subchapter_num)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_codes_classant ON ccam_codes(classant)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_assoc_code ON associations(code)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_assoc_associated ON associations(associated_code)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_codes_date_end ON ccam_codes(date_end)")
    conn.commit()
    print("  Indexes created.")


def print_stats(conn):
    """Print database statistics."""
    c = conn.cursor()
    print("\n=== Database Stats ===")
    c.execute("SELECT COUNT(*) FROM ccam_codes")
    print(f"  Total CCAM codes: {c.fetchone()[0]}")
    c.execute("SELECT COUNT(*) FROM ccam_codes WHERE date_end IS NULL")
    print(f"  Active codes (no end date): {c.fetchone()[0]}")
    c.execute("SELECT COUNT(*) FROM ccam_codes WHERE date_end IS NOT NULL")
    print(f"  Expired codes: {c.fetchone()[0]}")
    c.execute("SELECT COUNT(*) FROM ccam_codes WHERE icr_public IS NOT NULL")
    print(f"  Codes with ICR public: {c.fetchone()[0]}")
    c.execute("SELECT COUNT(*) FROM associations")
    print(f"  Total associations: {c.fetchone()[0]}")
    c.execute("SELECT association_type, COUNT(*) FROM associations GROUP BY association_type")
    for row in c.fetchall():
        print(f"    {row[0]}: {row[1]}")
    c.execute("SELECT COUNT(DISTINCT num) FROM chapters")
    print(f"  Total chapters/sections: {c.fetchone()[0]}")
    c.execute("SELECT COUNT(DISTINCT chapter_num) FROM ccam_codes WHERE chapter_num IS NOT NULL")
    print(f"  Distinct chapters used: {c.fetchone()[0]}")

    # Sample searches
    test_queries = [
        "arthrodèse cervicale",
        "craniotomie tumeur",
        "appendicectomie",
    ]
    for query in test_queries:
        norm_query = normalize_for_search(query)
        # FTS5: use implicit AND between terms
        print(f"\n=== Search: '{query}' (normalized: '{norm_query}') ===")
        try:
            c.execute("""
                SELECT ccam_codes.code, ccam_codes.label, ccam_codes.icr_public, ccam_codes.chapter_title
                FROM ccam_fts
                JOIN ccam_codes ON ccam_codes.rowid = ccam_fts.rowid
                WHERE ccam_fts.label_normalized MATCH ?
                ORDER BY ccam_fts.rank
                LIMIT 5
            """, (norm_query,))
            results = c.fetchall()
            if results:
                for row in results:
                    icr = f"ICR={row[2]}" if row[2] else "no ICR"
                    print(f"  {row[0]} | {row[1][:70]} | {icr}")
            else:
                print("  (no results)")
        except Exception as e:
            print(f"  [ERR] {e}")

    print(f"\n  Database size: {DB_PATH.stat().st_size / 1024:.0f} KB")


def main():
    print("T2A Assistant — CCAM Database Builder")
    print(f"Output: {DB_PATH}\n")

    conn = create_database()
    count = parse_complementary_file(conn)

    if count > 0:
        build_fts_index(conn)
        create_indexes(conn)
        print_stats(conn)
    else:
        print("[ERROR] No codes parsed. Check data files.")

    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
